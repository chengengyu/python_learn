__author__ = 'eling'
from openpyxl import Workbook, load_workbook
import copy

class BugInfo(object):
   def __init__(self, BugNum, OpenInfo, ResolveInfo, CloseInfo):
       self.BugNum = BugNum
       self.OpenInfo = OpenInfo
       self.ResolveInfo = ResolveInfo
       self.CloseInfo = CloseInfo


SourceFileName = input("上一次的汇总表xls: ")
DesFileName = input("本次的汇总表:")

SourWb = load_workbook(SourceFileName)
SourWs = SourWb.active
Sour = {}
#读取源表，讲内容根据DTMUC号存放在字典中
for row in SourWs.rows:
    bug = BugInfo(row[0].value, row[1].value, row[2].value, row[3].value)
    Sour[bug.BugNum] = bug

#读取新的表
DesWb = load_workbook(DesFileName)
DesWs = DesWb["高层算法组"]
SaveWb = Workbook()
SaveWs = SaveWb.active
SaveWs.title = "高层算法组"
#将算法组的内容拷贝到一个新的表中
for numRow, row in enumerate(DesWs.rows):
    SaveWs.append(row)
#遍历新的表，判断DTMUC号如果存在在源表中，则讲需要copy的内容复制过去
for num, rowEx in enumerate(SaveWs.rows):
    if rowEx[0].value in Sour:
        print(rowEx[0].value )
        bugSour = Sour[rowEx[0].value]
        #print(bugSour.OpenInfo)
        if bugSour.OpenInfo:
            cellNum = 'B'+str(num + 1)
            SaveWs[cellNum] = bugSour.OpenInfo
        if bugSour.ResolveInfo:
            cellNum = 'C'+str(num + 1)
            SaveWs[cellNum] = bugSour.ResolveInfo
        if bugSour.CloseInfo:
            cellNum = 'D'+str(num + 1)
            SaveWs[cellNum] = bugSour.CloseInfo
SaveWb.save("test.xlsx")
input("All Done, press any key to continue.")